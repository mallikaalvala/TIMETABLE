"""
B.Pharm Exam Timetable Generator
- Subjects from PCI B.Pharm Regulations 2014
- Morning 10 AM–1 PM | Afternoon 2 PM–5 PM
- No Sunday / Holiday exams
- ODD (I,III,V,VII) or EVEN (II,IV,VI,VIII) session
- Main + Backlog scheduling with clash detection
- Download as XLSX or PDF
"""

import streamlit as st
import pandas as pd
from datetime import date, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

# ─── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="📅 B.Pharm Exam Timetable", page_icon="📅",
                   layout="wide")

# ─── Custom CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
.title-box{background:linear-gradient(135deg,#1a237e,#283593);color:white;
  padding:18px 24px;border-radius:10px;margin-bottom:20px;text-align:center}
.section-head{background:#e8eaf6;padding:8px 14px;border-left:4px solid #3949ab;
  border-radius:4px;font-weight:bold;margin:12px 0 8px 0}
.clash-badge{background:#ffebee;color:#c62828;padding:4px 10px;
  border-radius:12px;font-size:12px;font-weight:bold}
.ok-badge{background:#e8f5e9;color:#2e7d32;padding:4px 10px;
  border-radius:12px;font-size:12px;font-weight:bold}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="title-box">
  <h2 style="margin:0">📅 B.Pharm Automatic Exam Timetable Generator</h2>
  <p style="margin:4px 0 0 0;opacity:0.85">
    G.Pulla Reddy College of Pharmacy 
  </p>
</div>
""", unsafe_allow_html=True)

# ─── Subject Database (from PCI PDF) ──────────────────────────────────────────
SUBJECTS = {
    "I": [
        ("BP101T", "Human Anatomy and Physiology I – Theory"),
        ("BP102T", "Pharmaceutical Analysis I – Theory"),
        ("BP103T", "Pharmaceutics I – Theory"),
        ("BP104T", "Pharmaceutical Inorganic Chemistry – Theory"),
        ("BP105T", "Communication Skills – Theory*"),
        ("BP107P", "Human Anatomy and Physiology – Practical"),
        ("BP108P", "Pharmaceutical Analysis I – Practical"),
        ("BP109P", "Pharmaceutics I – Practical"),
        ("BP110P", "Pharmaceutical Inorganic Chemistry – Practical"),
    ],
    "II": [
        ("BP201T", "Human Anatomy and Physiology II – Theory"),
        ("BP202T", "Pharmaceutical Organic Chemistry I – Theory"),
        ("BP203T", "Biochemistry – Theory"),
        ("BP204T", "Pathophysiology – Theory"),
        ("BP205T", "Computer Applications in Pharmacy – Theory*"),
        ("BP206T", "Environmental Sciences – Theory*"),
        ("BP207P", "Human Anatomy and Physiology II – Practical"),
        ("BP208P", "Pharmaceutical Organic Chemistry I – Practical"),
        ("BP209P", "Biochemistry – Practical"),
        ("BP210P", "Computer Applications in Pharmacy – Practical*"),
    ],
    "III": [
        ("BP301T", "Pharmaceutical Organic Chemistry II – Theory"),
        ("BP302T", "Physical Pharmaceutics I – Theory"),
        ("BP303T", "Pharmaceutical Microbiology – Theory"),
        ("BP304T", "Pharmaceutical Engineering – Theory"),
        ("BP305P", "Pharmaceutical Organic Chemistry II – Practical"),
        ("BP306P", "Physical Pharmaceutics I – Practical"),
        ("BP307P", "Pharmaceutical Microbiology – Practical"),
        ("BP308P", "Pharmaceutical Engineering – Practical"),
    ],
    "IV": [
        ("BP401T", "Pharmaceutical Organic Chemistry III – Theory"),
        ("BP402T", "Medicinal Chemistry I – Theory"),
        ("BP403T", "Physical Pharmaceutics II – Theory"),
        ("BP404T", "Pharmacology I – Theory"),
        ("BP405T", "Pharmacognosy and Phytochemistry I – Theory"),
        ("BP406P", "Medicinal Chemistry I – Practical"),
        ("BP407P", "Physical Pharmaceutics II – Practical"),
        ("BP408P", "Pharmacology I – Practical"),
        ("BP409P", "Pharmacognosy and Phytochemistry I – Practical"),
    ],
    "V": [
        ("BP501T", "Medicinal Chemistry II – Theory"),
        ("BP502T", "Industrial Pharmacy I – Theory"),
        ("BP503T", "Pharmacology II – Theory"),
        ("BP504T", "Pharmacognosy and Phytochemistry II – Theory"),
        ("BP505T", "Pharmaceutical Jurisprudence – Theory"),
        ("BP506P", "Industrial Pharmacy I – Practical"),
        ("BP507P", "Pharmacology II – Practical"),
        ("BP508P", "Pharmacognosy and Phytochemistry II – Practical"),
    ],
    "VI": [
        ("BP601T", "Medicinal Chemistry III – Theory"),
        ("BP602T", "Pharmacology III – Theory"),
        ("BP603T", "Herbal Drug Technology – Theory"),
        ("BP604T", "Biopharmaceutics and Pharmacokinetics – Theory"),
        ("BP605T", "Pharmaceutical Biotechnology – Theory"),
        ("BP606T", "Quality Assurance – Theory"),
        ("BP607P", "Medicinal Chemistry III – Practical"),
        ("BP608P", "Pharmacology III – Practical"),
        ("BP609P", "Herbal Drug Technology – Practical"),
    ],
    "VII": [
        ("BP701T", "Instrumental Methods of Analysis – Theory"),
        ("BP702T", "Industrial Pharmacy II – Theory"),
        ("BP703T", "Pharmacy Practice – Theory"),
        ("BP704T", "Novel Drug Delivery System – Theory"),
        ("BP705P", "Instrumental Methods of Analysis – Practical"),
        ("BP706PS", "Practice School"),
    ],
    "VIII": [
        ("BP801T", "Biostatistics and Research Methodology – Theory"),
        ("BP802T", "Social and Preventive Pharmacy – Theory"),
        ("BP803ET", "Pharmaceutical Marketing Management – Theory"),
        ("BP804ET", "Pharmaceutical Regulatory Science – Theory"),
        ("BP805ET", "Pharmacovigilance – Theory"),
        ("BP806ET", "Quality Control & Standardization of Herbals – Theory"),
        ("BP807ET", "Computer Aided Drug Design – Theory"),
        ("BP808ET", "Cell and Molecular Biology – Theory"),
        ("BP809ET", "Cosmetic Science – Theory"),
        ("BP810ET", "Experimental Pharmacology – Theory"),
        ("BP811ET", "Advanced Instrumentation Techniques – Theory"),
        ("BP812ET", "Dietary Supplements and Nutraceuticals – Theory"),
        ("BP813PW", "Project Work"),
    ],
}

ODD_SEMS  = ["I", "III", "V", "VII"]
EVEN_SEMS = ["II", "IV", "VI", "VIII"]

SLOT_LABELS = {
    "Morning":   "10:00 AM – 01:00 PM",
    "Afternoon": "02:00 PM – 05:00 PM",
}

# ─── Helper: available exam dates ─────────────────────────────────────────────
def get_exam_dates(start: date, end: date, holidays: list[date]) -> list[date]:
    dates = []
    d = start
    while d <= end:
        if d.weekday() != 6 and d not in holidays:   # 6 = Sunday
            dates.append(d)
        d += timedelta(days=1)
    return dates

# ─── Helper: build subject list for scheduling ───────────────────────────────
def build_subject_list(session: str, include_backlog: bool,
                        sel_main: list[str], sel_backlog: list[str]) -> list[dict]:
    """
    Returns ordered list of subjects to schedule.
    Strategy: interleave theory & practical, theory first.
    """
    main_sems = ODD_SEMS if session == "ODD" else EVEN_SEMS
    back_sems = EVEN_SEMS if session == "ODD" else ODD_SEMS

    rows = []
    # ── Main semesters ────────────────────────────────────────────────────────
    for sem in main_sems:
        if sem not in sel_main:
            continue
        for code, name in SUBJECTS[sem]:
            rows.append({
                "code": code, "subject": name,
                "semester": sem, "type": "Main",
                "slot": "", "date": None,
            })
    # ── Backlog semesters ──────────────────────────────────────────────────────
    if include_backlog:
        for sem in back_sems:
            if sem not in sel_backlog:
                continue
            for code, name in SUBJECTS[sem]:
                rows.append({
                    "code": code, "subject": name,
                    "semester": sem, "type": "Backlog",
                    "slot": "", "date": None,
                })
    return rows

# ─── Core: auto-assign dates + slots ──────────────────────────────────────────
def auto_schedule(subjects: list[dict], exam_dates: list[date]) -> pd.DataFrame:
    """
    Assign date + slot to each subject ensuring:
    - No two subjects that same student may attempt are in same slot/date.
    - Students in Sem N (main) may also have backlogs from lower sems of same parity.
    - Theory before Practical within each semester.
    """
    # Separate theory and practicals; theory first
    theory = [s for s in subjects if not s["code"].endswith("P")
              and not s["code"].endswith("PS") and not s["code"].endswith("PW")]
    practical = [s for s in subjects if s["code"].endswith("P")
                 or s["code"].endswith("PS") or s["code"].endswith("PW")]
    ordered = theory + practical

    # Track slot occupancy: (date, slot) → set of semesters already placed
    slot_occupancy: dict[tuple, set] = {}
    # Also track (date, slot) → list of subjects (for clash display)
    slot_items: dict[tuple, list] = {}

    scheduled = []
    date_idx   = 0
    slot_cycle = ["Morning", "Afternoon"]
    slot_idx   = 0

    for subj in ordered:
        placed = False
        attempts = 0
        while not placed and attempts < len(exam_dates) * 2:
            if date_idx >= len(exam_dates):
                break
            d    = exam_dates[date_idx]
            slot = slot_cycle[slot_idx]
            key  = (d, slot)

            occupied_sems = slot_occupancy.get(key, set())
            # Check: can we place this semester here?
            if subj["semester"] not in occupied_sems:
                slot_occupancy.setdefault(key, set()).add(subj["semester"])
                slot_items.setdefault(key, []).append(subj["code"])
                subj = dict(subj)
                subj["date"] = d
                subj["slot"] = slot
                subj["time"] = SLOT_LABELS[slot]
                scheduled.append(subj)
                placed = True

            # Advance slot/date
            slot_idx += 1
            if slot_idx >= len(slot_cycle):
                slot_idx = 0
                date_idx += 1
            attempts += 1

        if not placed:
            # Force place (out of dates) — mark as unscheduled
            subj = dict(subj)
            subj["date"] = None
            subj["slot"] = "⚠️ Unscheduled"
            subj["time"] = "—"
            scheduled.append(subj)

    df = pd.DataFrame(scheduled)
    if "date" in df.columns and df["date"].notna().any():
        df["date_str"] = df["date"].apply(
            lambda x: x.strftime("%d-%b-%Y (%a)") if pd.notna(x) and x else "—"
        )
    else:
        df["date_str"] = "—"
    return df

# ─── Clash Detector ───────────────────────────────────────────────────────────
def detect_clashes(df: pd.DataFrame) -> list[str]:
    clashes = []
    grouped = df.groupby(["date_str", "slot"])
    for (d, slot), grp in grouped:
        sems = grp["semester"].tolist()
        if len(sems) != len(set(sems)):
            dups = [s for s in set(sems) if sems.count(s) > 1]
            clashes.append(
                f"⚠️ {d} [{slot}] — Semester(s) {', '.join(dups)} has multiple subjects!"
            )
    return clashes

# ─── Excel Export ─────────────────────────────────────────────────────────────
def export_excel(df: pd.DataFrame, title: str) -> bytes:
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Exam Timetable"

    # Colour palette
    HDR_FILL  = PatternFill("solid", fgColor="1A237E")
    SEM_FILLS = {
        "I":    PatternFill("solid", fgColor="E3F2FD"),
        "II":   PatternFill("solid", fgColor="F3E5F5"),
        "III":  PatternFill("solid", fgColor="E8F5E9"),
        "IV":   PatternFill("solid", fgColor="FFF3E0"),
        "V":    PatternFill("solid", fgColor="FCE4EC"),
        "VI":   PatternFill("solid", fgColor="E0F7FA"),
        "VII":  PatternFill("solid", fgColor="F9FBE7"),
        "VIII": PatternFill("solid", fgColor="EDE7F6"),
    }
    MAIN_FILL = PatternFill("solid", fgColor="C8E6C9")
    BACK_FILL = PatternFill("solid", fgColor="FFECB3")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = HDR_FILL
    ws["A1"].alignment = center

    # Sub-header
    ws.merge_cells("A2:H2")
    ws["A2"] = "G.Pulla Reddy College of Pharmacy, Hyderabad | PCI B.Pharm Regulations 2014"
    ws["A2"].font = Font(italic=True, size=10, color="5C6BC0")
    ws["A2"].alignment = center

    # Column headers
    COLS = ["#", "Date", "Day", "Slot", "Time", "Semester", "Course Code",
            "Subject Name", "Type"]
    for ci, h in enumerate(COLS, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor="283593")
        cell.alignment = center
        cell.border    = border

    col_widths = [4, 18, 10, 12, 22, 10, 14, 48, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), 4):
        d_obj = row.get("date")
        day_name = d_obj.strftime("%A") if pd.notna(d_obj) and d_obj else "—"
        vals = [
            ri - 3,
            row.get("date_str", "—"),
            day_name,
            row.get("slot", ""),
            row.get("time", ""),
            f"Sem {row.get('semester','')}",
            row.get("code", ""),
            row.get("subject", ""),
            row.get("type", ""),
        ]
        fill = SEM_FILLS.get(row.get("semester", "I"),
                              PatternFill("solid", fgColor="FFFFFF"))
        type_fill = MAIN_FILL if row.get("type") == "Main" else BACK_FILL

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = border
            cell.font      = Font(size=9)
            cell.alignment = left if ci in (8,) else center
            if ci == 6:
                cell.fill = fill
            elif ci == 9:
                cell.fill = type_fill
            elif ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8F9FA")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 20
    for r in range(4, ws.max_row + 1):
        ws.row_dimensions[r].height = 18

    # Freeze panes
    ws.freeze_panes = "A4"

    # Legend sheet
    ls = wb.create_sheet("Legend & Rules")
    ls["A1"] = "Timetable Rules & Legend"
    ls["A1"].font = Font(bold=True, size=13, color="1A237E")
    rules = [
        ("Session Type",  "ODD = Sem I,III,V,VII Main | EVEN = Sem II,IV,VI,VIII Main"),
        ("Exam Timings",  "Morning: 10:00 AM – 01:00 PM | Afternoon: 02:00 PM – 05:00 PM"),
        ("Backlog",       "Students can write Main + Backlog; no two subjects in same slot"),
        ("Sundays",       "No exams on Sundays or declared holidays"),
        ("Clash Rule",    "Same semester never gets two different subjects in same slot"),
        ("* Subjects",    "Asterisk = Non-University Examination (NUE), conducted at college"),
    ]
    for i, (k, v) in enumerate(rules, 3):
        ls.cell(row=i, column=1, value=k).font = Font(bold=True, size=10)
        ls.cell(row=i, column=2, value=v).font  = Font(size=10)
    ls.column_dimensions["A"].width = 18
    ls.column_dimensions["B"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─── PDF Export ───────────────────────────────────────────────────────────────
def export_pdf(df: pd.DataFrame, title: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", fontSize=14, fontName="Helvetica-Bold",
                                  alignment=TA_CENTER, textColor=colors.HexColor("#1A237E"),
                                  spaceAfter=4)
    sub_style   = ParagraphStyle("s", fontSize=9,  fontName="Helvetica",
                                  alignment=TA_CENTER, textColor=colors.grey, spaceAfter=10)

    SEM_COLORS = {
        "I":    colors.HexColor("#E3F2FD"), "II":  colors.HexColor("#F3E5F5"),
        "III":  colors.HexColor("#E8F5E9"), "IV":  colors.HexColor("#FFF3E0"),
        "V":    colors.HexColor("#FCE4EC"), "VI":  colors.HexColor("#E0F7FA"),
        "VII":  colors.HexColor("#F9FBE7"), "VIII":colors.HexColor("#EDE7F6"),
    }

    story = [
        Paragraph(title, title_style),
        Paragraph("G.Pulla Reddy College of Pharmacy, Hyderabad — PCI B.Pharm Regulations 2014",
                  sub_style),
    ]

    # Table data
    header = ["#", "Date", "Day", "Slot", "Time", "Sem", "Code", "Subject", "Type"]
    data   = [header]
    row_colors = [colors.HexColor("#1A237E")]  # header bg

    for idx, (_, row) in enumerate(df.iterrows(), 1):
        d_obj = row.get("date")
        day_n = d_obj.strftime("%a") if pd.notna(d_obj) and d_obj else "—"
        data.append([
            str(idx),
            row.get("date_str", "—"),
            day_n,
            row.get("slot", ""),
            row.get("time", ""),
            f"Sem {row.get('semester','')}",
            row.get("code", ""),
            row.get("subject", ""),
            row.get("type", ""),
        ])
        sem = row.get("semester", "I")
        bg  = SEM_COLORS.get(sem, colors.white)
        row_colors.append(bg if idx % 2 == 0 else colors.white)

    col_widths = [1*cm, 3.2*cm, 1.6*cm, 2.2*cm, 3.8*cm,
                  1.5*cm, 2.4*cm, 9*cm, 1.8*cm]

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#1A237E")),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0), 8),
        ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",        (7, 1), (7, -1), "LEFT"),
        ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",     (0, 1), (-1, -1), 7.5),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F5F5F5")]),
        ("GRID",         (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
    ]
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)

    # Footer note
    story.append(Spacer(1, 0.4*cm))
    note_style = ParagraphStyle("n", fontSize=7, fontName="Helvetica",
                                 textColor=colors.grey)
    story.append(Paragraph(
        "* NUE = Non-University Examination (conducted at college level). "
        "Morning: 10:00 AM–1:00 PM | Afternoon: 2:00 PM–5:00 PM. "
        "No exams on Sundays or declared holidays.", note_style))

    doc.build(story)
    return buf.getvalue()

# ═══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ═══════════════════════════════════════════════════════════════════════════════

# ─── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Configuration")
    st.divider()

    session = st.radio("📚 Exam Session",
                       ["ODD (I, III, V, VII)", "EVEN (II, IV, VI, VIII)"],
                       help="ODD semesters: Nov/Dec | EVEN semesters: May/Jun")
    session_type = "ODD" if session.startswith("ODD") else "EVEN"
    main_sems = ODD_SEMS if session_type == "ODD" else EVEN_SEMS
    back_sems = EVEN_SEMS if session_type == "ODD" else ODD_SEMS

    st.divider()
    st.markdown("### 📅 Date Range")
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start Date",
                                    value=date(2025, 11, 3), min_value=date(2020,1,1))
    with col2:
        end_date   = st.date_input("End Date",
                                    value=date(2025, 12, 20), min_value=date(2020,1,1))

    st.divider()
    st.markdown("### 🏖️ Holidays (skip these dates)")
    holiday_input = st.text_area(
        "One date per line (DD-MM-YYYY)",
        placeholder="03-11-2025\n14-11-2025\n25-12-2025",
        height=100,
    )
    holidays = []
    for line in holiday_input.strip().splitlines():
        line = line.strip()
        if line:
            try:
                holidays.append(
                    date(*reversed([int(x) for x in line.split("-")]))
                )
            except Exception:
                st.warning(f"Invalid date: {line}")

    st.divider()
    st.markdown("### 📖 Select Semesters")
    sel_main = st.multiselect(
        "Main Semesters", main_sems, default=main_sems,
        help=f"Current {session_type} batch semesters"
    )
    include_backlog = st.checkbox("Include Backlog Semesters", value=True)
    sel_backlog = []
    if include_backlog:
        sel_backlog = st.multiselect(
            "Backlog Semesters", back_sems, default=back_sems,
            help="Students who failed previous sessions"
        )

    st.divider()
    col_title = st.text_input("📄 Report Title",
                               value="B.Pharm End Semester Examination Timetable")

# ─── Main Panel ───────────────────────────────────────────────────────────────
if not sel_main and not sel_backlog:
    st.info("👈 Select at least one semester from the sidebar to generate a timetable.")
    st.stop()

if start_date >= end_date:
    st.error("End date must be after start date.")
    st.stop()

exam_dates = get_exam_dates(start_date, end_date, holidays)
avail_slots = len(exam_dates) * 2

subjects = build_subject_list(session_type, include_backlog,
                               sel_main, sel_backlog)
total_subj = len(subjects)

# ─── Stats banner ─────────────────────────────────────────────────────────────
c1, c2, c3, c4 = st.columns(4)
c1.metric("📚 Total Subjects", total_subj)
c2.metric("📅 Exam Days", len(exam_dates))
c3.metric("🕐 Available Slots", avail_slots)
c4.metric("✅ Coverage",
          "✔ OK" if avail_slots >= total_subj else "⚠️ Need more days")

st.divider()

# ─── Generate Button ──────────────────────────────────────────────────────────
if st.button("🚀 Generate Timetable", type="primary", use_container_width=True):
    if avail_slots < total_subj:
        st.warning(f"⚠️ Only {avail_slots} slots available for {total_subj} subjects. "
                   f"Some subjects may be unscheduled. Extend the date range.")

    df = auto_schedule(subjects, exam_dates)
    st.session_state["df"]    = df
    st.session_state["title"] = col_title
    st.success(f"✅ Timetable generated for **{total_subj}** subjects "
               f"across **{len(exam_dates)}** exam days!")

# ─── Display & Edit ───────────────────────────────────────────────────────────
if "df" in st.session_state:
    df    = st.session_state["df"]
    title = st.session_state["title"]

    # ── Clash Check ───────────────────────────────────────────────────────────
    clashes = detect_clashes(df)
    if clashes:
        st.error("### ⚠️ Clash Detected!")
        for c in clashes:
            st.markdown(f"- {c}")
    else:
        st.success("✅ No clashes detected — all semesters have distinct time slots.")

    st.divider()

    # ── Semester Filter ────────────────────────────────────────────────────────
    st.markdown('<div class="section-head">📋 Timetable Preview & Manual Edit</div>',
                unsafe_allow_html=True)

    filter_sems = st.multiselect(
        "Filter by Semester",
        options=sorted(df["semester"].unique()),
        default=sorted(df["semester"].unique()),
        key="filter_sems",
    )
    view_df = df[df["semester"].isin(filter_sems)].copy()

    # Display columns
    disp_cols = ["date_str", "slot", "time", "semester", "type", "code", "subject"]
    rename_map = {
        "date_str": "Date", "slot": "Slot", "time": "Time",
        "semester": "Semester", "type": "Type",
        "code": "Course Code", "subject": "Subject",
    }

    edited = st.data_editor(
        view_df[disp_cols].rename(columns=rename_map),
        use_container_width=True,
        num_rows="fixed",
        column_config={
            "Date":        st.column_config.TextColumn("Date", width="medium"),
            "Slot":        st.column_config.SelectboxColumn(
                               "Slot", options=["Morning", "Afternoon"], width="small"),
            "Time":        st.column_config.TextColumn("Time", width="medium"),
            "Semester":    st.column_config.TextColumn("Semester", width="small"),
            "Type":        st.column_config.SelectboxColumn(
                               "Type", options=["Main", "Backlog"], width="small"),
            "Course Code": st.column_config.TextColumn("Course Code", width="small"),
            "Subject":     st.column_config.TextColumn("Subject", width="large"),
        },
        key="editor",
        hide_index=True,
    )

    # Re-check clashes after edits
    edited_back = edited.rename(columns={v: k for k, v in rename_map.items()})
    for col in ["date", "semester", "type", "code"]:
        if col not in edited_back.columns:
            edited_back[col] = view_df[col].values
    post_clashes = detect_clashes(edited_back)
    if post_clashes:
        st.error("### 🔴 Manual Edit Introduced Clashes!")
        for c in post_clashes:
            st.markdown(f"- {c}")
        st.warning("Please resolve the clashes before downloading.")
    else:
        if len(edited) < len(view_df) or any(
            edited.iloc[i]["Slot"] != view_df.iloc[i]["slot"]
            for i in range(min(len(edited), len(view_df)))
        ):
            st.success("✅ Manual edits look good — no clashes.")

    # ── Summary per Semester ──────────────────────────────────────────────────
    st.divider()
    st.markdown('<div class="section-head">📊 Summary by Semester</div>',
                unsafe_allow_html=True)
    summary = (df.groupby(["semester", "type"])
                 .size().reset_index(name="count")
                 .pivot(index="semester", columns="type", values="count")
                 .fillna(0).astype(int))
    summary.index.name = "Semester"
    st.dataframe(summary, use_container_width=True)

    # ── Date-wise View ────────────────────────────────────────────────────────
    st.divider()
    st.markdown('<div class="section-head">🗓️ Day-wise Schedule</div>',
                unsafe_allow_html=True)
    pivot = (df[df["date"].notna()]
             .groupby(["date_str", "slot"])["subject"]
             .apply(lambda x: " | ".join(
                 [f"Sem {row['semester']} ({row['type'][0]}): {row['subject']}"
                  for _, row in df[
                      (df["date_str"] == x.name[0]) & (df["slot"] == x.name[1])
                  ].iterrows()]))
             .unstack(fill_value="—")
             .reset_index()
             .rename(columns={"date_str": "Date"}))
    st.dataframe(pivot, use_container_width=True, hide_index=True)

    # ── Downloads ─────────────────────────────────────────────────────────────
    st.divider()
    st.markdown('<div class="section-head">⬇️ Download Timetable</div>',
                unsafe_allow_html=True)
    dcol1, dcol2 = st.columns(2)

    with dcol1:
        xlsx_bytes = export_excel(df, title)
        st.download_button(
            label="📊 Download as Excel (.xlsx)",
            data=xlsx_bytes,
            file_name="BPharm_ExamTimetable.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

    with dcol2:
        pdf_bytes = export_pdf(df, title)
        st.download_button(
            label="📄 Download as PDF",
            data=pdf_bytes,
            file_name="BPharm_ExamTimetable.pdf",
            mime="application/pdf",
            use_container_width=True,
            type="primary",
        )

    st.caption(
        "💡 Tip: Edit any cell in the timetable above. "
        "If a change causes a clash (same semester in two slots on same day), "
        "a red warning will appear automatically."
    )

else:
    st.info("👆 Click **Generate Timetable** to auto-schedule all subjects.")
    with st.expander("📖 Subject List (from PCI PDF)"):
        for sem, subs in SUBJECTS.items():
            st.markdown(f"**Semester {sem}** ({len(subs)} subjects)")
            for code, name in subs:
                st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;`{code}` — {name}")
